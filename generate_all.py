import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
import selenium
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
import os.path
from selenium import webdriver
import time
import glob, os
import json
import os
import openpyxl
import re
# import pinyin
from pypinyin import pinyin, lazy_pinyin, Style
import shutil
from PIL import Image
import selenium
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
import os.path
from selenium import webdriver

from pathlib import Path
from pandas.api.types import is_string_dtype
from bs4 import BeautifulSoup
from pandas.api.types import is_numeric_dtype
from deep_translator import (GoogleTranslator,
                             MicrosoftTranslator,
                             PonsTranslator,
                             LingueeTranslator,
                             MyMemoryTranslator,
                             YandexTranslator,
                             PapagoTranslator,
                             DeeplTranslator,
                             QcriTranslator,
                             single_detection,
                             batch_detection)

link = {}


# this does not cleanup
def get_raw_dict_df():
    for excel_data_raw in glob.glob("*.xlsx"):
        if excel_data_raw != "class_map.xlsx":
            dict_df = pd.read_excel(excel_data_raw, sheet_name=None)
            wb = openpyxl.load_workbook(excel_data_raw)
            dict_df_wb_pair = [dict_df, wb]
            return dict_df_wb_pair


dict_df_wb_pair = []  # to accelerate replace, not repeating the process


def get_dict_df():
    global dict_df_wb_pair
    if dict_df_wb_pair:
        return dict_df_wb_pair

    for excel_data_raw in glob.glob("*.xlsx"):
        if excel_data_raw != "class_map.xlsx":
            dict_df = cleanup(pd.read_excel(excel_data_raw, sheet_name=None, engine="openpyxl"),
                              excel_data_raw=excel_data_raw)
            wb = openpyxl.load_workbook(excel_data_raw)
            dict_df_wb_pair = [dict_df, wb]
            return dict_df_wb_pair


def space_replacer(word):
    if isinstance(word, str):
        return word.replace(" ", "_")
    else:
        return word


def underscore_replacer(word):
    return word.replace("_", " ")


def generate_lark_to_mc2_link():
    dict_df, wb = get_dict_df()
    for tab, df in dict_df.items():
        print(
            f'HYPERLINK(CONCATENATE("https://unforgiven1990.github.io/mc2/page/","{tab}","/",SUBSTITUTE([{tab}]," ","_"),".html"), "LINK")')


def sort_dict_by_key_length(d):
    new_d = {}
    for k in sorted(d, key=len, reverse=True):
        new_d[k] = d[k]

    return new_d

def maketitle(string):
    return ' '.join([w.title() if w.islower() else w for w in string.split()])


def cleanup(dict_df, excel_data_raw):
    """
    iterate through all tabs, changes key items with comma to other sign
    1. copy each index as a copy column
    """
    new_dict_df = {}
    dict_to_replace = {}
    forbidden_chars = {"@nio.com": '',
                       "@nio.io": '',
                       "    ": '',
                       ",": '',
                       ".": '_',
                       '"': '',
                       "/": '',
                       ":": '',
                       ")":'',
                       "(":'',
                       "  ": ' ',  # replaced in this order, first remove double space, then replace to underscore
                       " ": '_',  # still empty space in tab data not converted
                       }

    dict_bugged = {  # todo bugged doesn't work
        "UserDevelopment": 'User Development',
        "UserOperation": "User Operation",
        "UserRelationship": "User Relationship",
        "UserRelation": "User Relation",
        "UserTeam": "User Team",
    }
    forbidden_chars = {**forbidden_chars}

    #Remove before proceed: remove leading and ending space for each index
    d_helper={}
    for tab, df in dict_df.items():
        df = df.replace(r"^ +| +$", r"", regex=True)
        d_helper[tab] = df.replace('\xa0', '', regex=True)
    dict_df=d_helper


    # for each tab in df- replace index
    for tab, df in dict_df.items():


        # create a new helper column to replace index
        df["RemoveMe"] = df[tab].copy()

        # copied_index_item is potentially wrong and needs to be replaced
        for copied_index_item, (index, row) in zip(df["RemoveMe"], df.iterrows()):
            # if index is none
            if pd.isna(copied_index_item):
                continue

            # replace all index data = index
            correct_item = copied_index_item  # starting position
            for forbidden_char, toreplace in forbidden_chars.items():
                try:
                    # correct_item = " ".join(correct_item.split()) #doing this removes double space in titles
                    correct_item = correct_item.replace(forbidden_char, toreplace)
                except:
                    correct_item = correct_item
            df.at[index, "RemoveMe"] = fr"{correct_item}"
            dict_to_replace[copied_index_item] = fr"{correct_item}"

        # actually replace index
        df[tab] = df["RemoveMe"]
        df.drop('RemoveMe', axis=1, inplace=True)


    # complete df replace why do you even need it?
    dict_to_replace=sort_dict_by_key_length(dict_to_replace)
    print(dict_to_replace)
    for tab, df in dict_df.items():
        if tab in ["Employee_Process","Capability"]:
            df.to_excel(f"w_before replace  {tab}.xlsx")

        # replace all in one
        if True:
            df=df.replace(dict_to_replace, regex=True)

            if tab in ["Employee_Process","Capability"]:
                df.to_excel(f"w_after replace  {tab}.xlsx")
        else:
            # replace row by row
            for col in df.columns:  # for each column
                for key, entirecell in df[col].items():  # for each index, value pair
                    try:
                        a_val = entirecell.splt(",")
                    except:
                        # value is not string, nothing to replace
                        continue

                    a_val_replaced = []
                    for val in a_val:  # for each a b c in "a,b,c"
                        # multiple values
                        if val in dict_to_replace:
                            a_val_replaced += [dict_to_replace[val]]
                    else:
                        df.at[key, entirecell] = ",".join(a_val_replaced)

        new_dict_df[tab] = df

    # replace related document link hyperlink with real title
    for tab in ["Employee_Process", "User_Process", ]:
        df_tab = new_dict_df[tab]

        # open the data with another library to get hyperlink
        wb = openpyxl.load_workbook(excel_data_raw)
        ws = wb[tab]

        for column in ["Related Document"]:
            column_index = df_tab.columns.get_loc(column)
            for index, cell_data in df_tab[column].items():
                if pd.notna(cell_data) and cell_data is not None:
                    try:
                        hyperlink = ws.cell(row=index + 2, column=column_index + 1).hyperlink.target
                    except:
                        hyperlink = cell_data

                    print(cell_data, hyperlink)
                    s_hyperlinks = ''
                    for link in hyperlink.split(" "):
                        hyperlink_display = hyperlink.replace("https://", "")
                        s_hyperlinks += f"<a href='{link}' target='_blank'>{hyperlink_display}</a>"

                    df_tab.at[index, column] = s_hyperlinks


    return new_dict_df


def return_string_gallery(word):
    """returns a fa icon"""
    dict_gallery = {
        "Department": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblKZKd1q8wmv6HL&view=vewUx4CSfi",
        "Employee": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblPtBnAeQA82JcL&view=vewbVi02lT",
        "Role": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblvaOn65LfwKZgh&view=vewtB3stx3",
        "User_Process": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblXI11nUP5hraec&view=vewdO1XHOV",
        "Employee_Process": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblmqm1WaKREMzM6&view=vew18ykPxw",
        "Capability": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblaQfACFL2RmmF0&view=vewyHEANB6",
        "System": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblkrJ6NVdCbc10a&view=vewmLZijxy",
        "Strategy": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbleijKoxmfs8WXB&view=vewZfdR1ir",
        "City": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblVk7GzlSl7A3fK&view=vewppJl3ro",
        "Topic": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblj5cHr1kISQ7C0&view=vewJsgQtHI",
        "KnowHow": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbl2nnVvMNfVLxWB&view=vewxRkGmvV",
        "Car": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblIjUl8dRCbolCI&view=vewEEtBGbz",
        "KPI": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblKMxqrBjWmaw7f&view=vew2GwzJXf",
        "Business_Model": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbl8tg5FsGjSKG3Y&view=vewzVWmrso",
        "Department_Category": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbl8tg5FsGjSKG3Y&view=vewzVWmrso",
        "User_Journey": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblAyuTuEKVVVrZ3&view=vewBZ5BYQK",
        "Process_Category": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblgDQkj9F2O3XSd&view=vewxV21Rjf",
        "Country": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblDmXtQ5JPySoUz&view=vewrnTjfGP",
        "Things": "fa-gear",
        "Facility": "fa-store",
        "Approval": "fa-check",
    }
    return dict_gallery[word]


def return_string_indirect(word):
    """returns a fa icon"""
    dict_gallery = {
        "Departments": [],
        "Department": ["Role", "City", "Capability", "Process_Category",  "Employee_Process", "System"],
        "Department_Category": ["Role", "Employee", "City", "Capability", "Process_Category",
                                "Employee_Process", "System"],
        "Employee": [ "Employee_Process", "Capability", "System"],
        "Role": [ "Department", "Capability", "System", "City"],
        "User_Process": ["Employee_Process", "Department", "Role", "Employee"],
        "Employee_Process": ["Department", "Department_Category",  "Employee", "Topic"],
        "Capability": ["Department", "Department_Category", "Role", "Employee"],
        "System": ["Department", "Department_Category", "Role", "Employee"],
        "Strategy": ["Employee_Process", "System"],
        "City": ["Role", "Department", "Employee_Process"],
        "Car": ["City"],
        "Business_Model": ["Department", "Department_Category", "Role", "Employee"],
        "User_Journey": ["Department", "Department_Category", "Role", "Employee"],
        "Process_Category": ["Department", "Department_Category", "Role", "Employee", "System"],
        "Country": ["Employee_Process"],
        "Things": [],
        "Facility": ["Country"],
        "Topic": ["Employee_Process"],
    }
    return dict_gallery[word]


def return_string_icon(word):
    """returns a fa icon"""
    dict_icon = {
        "Departments": "fa-sitemap",
        "L1_Department": "fa-folder-tree",
        "Department": "fa-folder-tree",
        "Department_Category": "fa-sitemap",
        "L2_Department": "fa-folder-tree",
        "L3_Department": "fa-folder-tree",
        "Leader": "fa-user-tie",
        "Employee": "fa-user",
        "Role": "fa-user-tag",
        "Process": "fa-repeat",
        "User_Process": "fa-left-right",
        "Employee_Process": "fa-people-arrows",
        "Capability": "fa-location-crosshairs",
        "System": "fa-screwdriver-wrench",
        "Value": "fa-hand-holding-heart",
        "People": "fa-users",
        "Strategy": "fa-compass",
        "City": "fa-location-dot",
        "Topic": "fa-chart-simple",
        "KnowHow": "fa-lightbulb",
        "Car": "fa-car",
        "KPI": "fa-chart-simple",
        "Business_Model": "fa-money-bill",
        "User_Journey": "fa-route",
        "Process_Category": "fa-list-check",
        "Country": "fa-globe",
        "Things": "fa-gear",
        "Approval": "fa-check",
        "Facility": "fa-store",
    }
    return dict_icon[word]


def return_string_component(word):
    """returns a fa icon"""
    dict_explainer = {
        "Departments": "",
        "L1_Department": "This view shows you what L1 department exists that are relevant for EB.",
        "Department": "This view shows you what L1 department exists that are relevant for EB.",
        "Department_Category": "This view is to show high level abstrac classes aggregated by many departments",
        "L2_Department": "This view shows you what L2 departments under European Business.",
        "L3_Department": "This view shows you what L3 departments under European Business.",
        "Leader": "This summary shows who are the department leaders and what do they lead.",
        "Employee": "This summary shows all employee details in European Business.",
        "Role": "This view shows the abstract role in European Business and who is working as them.",
        "Process": "This view shows what User_Process is there and what Employee_Process is there.",
        "User_Process": "This view shows what User_Process exists and how are they defined.",
        "Employee_Process": "This view shows what Employee_Process exists and how are they defined.",
        "Capability": "This view shows what Business Capabilities exists and what processes implements these capabilities",
        "System": "This view lists all relevant systems for European Business, how to use them and where to access them.",
        "Value": "This view lists all companies values and how it is reflected in our Business Capabilties.",
        "People": "asd",
        "KPI": "This view lists all relevant KPIs for their processes.",
        "Strategy": "This view shows what high level Strategies exists and which Processes implements these Strategies.",
        "City": "The location view shows what infrastructure is there and which employee is here.",
        "Topic": "The location view shows what infrastructure is there and which employee is here.",
        "KnowHow": "What Other Companies are doing, how NIO china is doing, what experiences we have learned",
        "Car": "What model exists for different country and business model",
        "Business_Model": "The location view shows what form of ownership user can have.",
        #"User_Journey": f"<select id='select_business'></select><select id='select_perspective'></select>",
        "User_Journey": f"",
        #"Process_Category": "<select id='select_business'></select><select id='select_perspective'></select>",
        "Process_Category": "",
        "Country": "Markets where NIO sells car",
        "Things": "sad",
        "Approval": "sad",
        "Facility": "sad",
    }
    return dict_explainer[word]


def return_string_editurl(word):
    """returns a fa icon"""
    dict_url = {
        "Department": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblyxOzBlxXbfgFi&view=vewgFkOi9f",
        "Department_Category": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbltmeBARSbQIJxN&view=vewN8BQVGq",
        "L1_Department": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblyxOzBlxXbfgFi&view=vewgFkOi9f",
        "L2_Department": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblcnQTN78GEt3nR&view=vewjua7iRe",
        "L3_Department": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblyxoLGbBcp1yUZ&view=vew8hBjN9a",
        "Leader": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblPFljTQ27fSZnc&view=vewNiWsvzO",
        "Employee": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblPtBnAeQA82JcL&view=vewM4stVK8",
        "Role": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblvaOn65LfwKZgh&view=vewwt7M9id",
        "User_Process": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblXI11nUP5hraec&view=vewX0IOZ4B",
        "Employee_Process": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblmqm1WaKREMzM6&view=vewAvMXYfY",
        "Capability": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblaQfACFL2RmmF0&view=vewKr0PzLw",
        "System": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblkrJ6NVdCbc10a&view=vewmLZijxy",
        "Value": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblIoyKb5UMTmu09&view=vewRvuZf3B",
        "Strategy": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbleijKoxmfs8WXB&view=vewj6vLShZ",
        "City": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblVk7GzlSl7A3fK&view=vewppJl3ro",
        "Topic": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblj5cHr1kISQ7C0&view=vewJsgQtHI",
        "KnowHow": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbl2nnVvMNfVLxWB&view=vewYJshw2X",
        "Car": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblIjUl8dRCbolCI&view=vewYO6qS8t",
        "KPI": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblKMxqrBjWmaw7f&view=vewAteOLwS",
        "Business_Model": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbl8tg5FsGjSKG3Y&view=vewBKzxKpU",
        "User_Journey": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblAyuTuEKVVVrZ3&view=vew6pEQzqw",
        "Process_Category": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblgDQkj9F2O3XSd&view=vewpiLPhLI",
        "Country": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblDmXtQ5JPySoUz&view=vewSKuK63B",
        "Approval": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tblLJfEXlvTIHzCz&view=vewgsFXq2c",
        "Facility": "https://nio.feishu.cn/wiki/wikcnHtTpp2T1YilHB3jiT3tiLf?table=tbl27IlPlyMwoUQ9&view=vewzwoeVMm",
    }
    return dict_url[word]


def return_global_navbar():
    dict_nav = {
        "People": ["Employee", "Role", "Department", "Department_Category"],
        "Strategy": ["Topic", "Strategy", "Capability", "Business_Model"],
        "Process": ["User_Journey", "User_Process", "Process_Category", "Employee_Process"],
        "Things": ["Country", "City", "Car", "Facility", "System"],
    }

    navbar_template = '<nav class="navbar navbar-expand-lg navbar-light bg-light" aria-label="Eighth navbar example"> <div class="container">  <a href="../../page/index/index.html" class="navbar-brand"> <img src="../../img/nio light.png" height="28" alt="CoolBrand"> </a>  <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarinstance" aria-controls="navbarinstance" aria-expanded="false" aria-label="Toggle navigation"> <span class="navbar-toggler-icon"></span> </button> <div class="collapse navbar-collapse" id="navbarinstance"> <ul class="navbar-nav me-auto mb-2 mb-lg-0"> {} </ul> </div> </div> </nav>'
    navbar_content = ""  # for li in url template
    for key, array in dict_nav.items():
        navbar_content_li_template = f'<li class="nav-item dropdown"> <a class="nav-link dropdown-toggle" href="#" id="navbarDropdown" role="button" data-bs-toggle="dropdown" aria-haspopup="true" aria-expanded="false"> <i class="fa-solid {return_string_icon(key)}"></i> {key} </a> <div class="dropdown-menu" aria-labelledby="navbarDropdown">' + '{}</div></li>'
        navbar_content_li_content = ""
        for counter, label in enumerate(array):
            navbar_content_li_content = navbar_content_li_content + f'<a class="dropdown-item" href="../../page/{label}/{label}.html"><i class="fa-solid {return_string_icon(label)}"></i> {underscore_replacer(label)}</a>'
            if counter + 1 < len(array):  # add divider
                navbar_content_li_content = navbar_content_li_content + f'<div class="dropdown-divider"></div>'

        # add new li to previous li
        navbar_content = navbar_content + navbar_content_li_template.format(navbar_content_li_content)

    global global_navbar
    global_navbar = navbar_template.format(navbar_content)


def return_global_html():
    # create data
    dict_df, wb = get_dict_df()

    # create json df by df
    all_json = ''
    for tab, df in dict_df.items():
        df.to_excel(f"helper/{tab}.xlsx")
        one_df = df.to_json(orient="records", lines=False, force_ascii=False, compression=None)
        one_df = f'"{tab}"' + ": " + one_df + ","
        all_json = all_json + one_df

    all_json = ' var data={' + all_json + "};"

    # bug is caused by pyhon df to_json function
    """bug is that some
    1. some department in bitable not linked 
    2. in bitable: user operation, in json useroperation
    when transition from df to json it klebs zusammen to Head of UserOperation Netherlands
    """

    dict_bugged = {
        "UserDevelopment": 'User Development',
        "UserOperation": "User Operation",
        "UserRelationship": "User Relationship",
        "UserRelation": "User Relation",
        "UserTeam": "User Team",
    }
    dict_bugged = {}
    for key, val in dict_bugged.items():
        all_json = all_json.replace(key, val)
    with open(fr"bootstrap/js/data.js", "w", encoding="utf-8") as file:
        file.write(str(all_json))

    # create edgematrix
    # <script src="https://unpkg.com/cytoscape/dist/cytoscape.min.js"></script>
    # <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.4.0/jquery.min.js"></script>

    test = """
        <!-- Cytoscape extention -->

		<!-- Jquery for graphml -->

    """

    local_extension = """
		<!-- Other iVis-at-Bilkent libraries -->
		<script src="https://unpkg.com/layout-base@1.0.2/layout-base.js"></script>
		<script src="https://unpkg.com/avsdf-base/avsdf-base.js"></script>
		<script src="https://unpkg.com/cose-base@1.0.3/cose-base.js"></script>
		<script src="https://unpkg.com/cytoscape-graphml/cytoscape-graphml.js"></script>
		<script src="https://unpkg.com/dagre@0.7.4/dist/dagre.js"></script>
	"""
    local_extension = """
    		<script  src="../../bootstrap/extensions/layout-base.js"></script>
    		<script  src="../../bootstrap/extensions/avsdf-base.js"></script>
    		<script  src="../../bootstrap/extensions/cose-base.js"></script>
    		<script  src="../../bootstrap/extensions/cytoscape-graphml.js"></script>
    		<script  src="../../bootstrap/extensions/dagre.js"></script>
    	"""

    test = test + local_extension

    meta = '<meta charset="utf-8"> <meta name="viewport" content="width=device-width, initial-scale=1">'
    js_jquery = '<script src="https://code.jquery.com/jquery-3.3.1.min.js" crossorigin="anonymous"></script>'
    js_popperjs = '<script src="https://cdn.jsdelivr.net/npm/popper.js@1.14.7/dist/umd/popper.min.js" integrity="sha384-UO2eT0CpHqdSJQ6hJty5KVphtPhzWj9WO1clHTMGa3JDZwrnQq4sF86dIHNDz0W1" crossorigin="anonymous"></script>'
    js_bootstrap = '<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.2.3/dist/js/bootstrap.bundle.min.js" integrity="sha384-kenU1KFdBIe4zVF0s0G1M5b4hcpxyD9F7jL+jjXkk+Q2h455rYXK/7HAuoJl+0I4" crossorigin="anonymous"></script>'
    js_fa = '<script defer src="https://use.fontawesome.com/releases/v5.15.4/js/all.js" integrity="sha384-rOA1PnstxnOBLzCLMcre8ybwbTmemjzdNlILg8O7z1lUkLXozs4DHonlDtnE7fpc" crossorigin="anonymous"></script>'
    js_cytoscape = '<script src="https://cdnjs.cloudflare.com/ajax/libs/cytoscape/3.23.0/cytoscape.min.js" integrity="sha512-gEWKnYYa1/1c3jOuT9PR7NxiVI1bwn02DeJGsl+lMVQ1fWMNvtjkjxIApTdbJ/wcDjQmbf+McWahXwipdC9bGA==" crossorigin="anonymous" referrerpolicy="no-referrer"></script>'
    js_cj = '<script  src="../../bootstrap/js/cj.js"  ></script>' + test
    js_cise = '<script  src="../../bootstrap/js/cytoscape-cise.js"></script>'
    js_dagre = '<script  src="../../bootstrap/js/cytoscape-dagre.js"></script>'
    js_data = js_dagre + js_cise + '<script  src="../../bootstrap/js/data.js"  ></script>'
    js_link = '<script  src="../../bootstrap/js/link.js"  ></script>'

    css_fa = '<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.2.1/css/all.min.css">'
    css_bootstrap = '<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.2.3/dist/css/bootstrap.min.css">'
    css_cj = '<link href="../../bootstrap/css/cj.css" rel="stylesheet">'

    favicon = '<link rel="icon" href="../../img/nio.ico">'

    head = '<head data-bs-spy="scroll" data-bs-target="#navbar-example" >{}</head>'.format(
        css_fa + css_bootstrap + css_cj + favicon)
    body = "<body>" + global_navbar + "<div class='container'>{content}</div> " + meta + js_jquery + js_popperjs + js_bootstrap + js_fa + js_cytoscape + "{jsinclude}" + js_cj + js_data + js_link + "</body>"
    bottomspacer = '<div class="p-2 m-2"></div>'
    footer = '<footer class="py-0 my-0 fixed-bottom"><p class="text-center text-disable px-3" style="float:right;">&copy; 2023 made by CJ</p></footer>'

    global template
    template = head + body + bottomspacer


# get all images of a particular instance
def return_instance_img(instance, tab, imgclass):
    a_img = []
    instance_with_space = instance.replace("_", " ")
    a_multiple = []
    a_multiple += [f"({x})" for x in range(10)]
    a_pic = [".jpg", ".png", ".webp",".svg",".jpeg",".JPG", ".PNG", ".WEBP",".SVG",".JPEG" ]
    a_pdf = [".pdf",".PDF"]

    for ending in a_pic + a_pdf:
        for file_path in glob.glob(f"attachments/Process Standardization Input-{tab}_Attachment/*{ending}"):
            for multiple in ["", "(1)", "(2)", "(3)", "(4)"]:
                file_name = os.path.basename(file_path).replace(f"{ending}", "").replace(f"_Process Flow{multiple}", "")
                # check this name for first image against this instance
                if instance_with_space == file_name:
                    url = f'../../attachments/Process Standardization Input-{tab}_Attachment/{instance.replace("_", " ")}_Process Flow{multiple}{ending}'
                    if ending in a_pic:
                        a_img += [f'<img class="{imgclass} mb-3" src="{url}" >']
                    elif ending in a_pdf:
                        a_img += [f"<embed src='{url}#toolbar=0'  style='height:90vh;  width:100%;' />"]
                    else:
                        a_img += [f'<img class="{imgclass} mb-3" src="{url}" >']

                # check this name for other images (n) against this instance
                if False:
                    for potential_counter in range(3):
                        # if file_name[-1]==")" and file_name[-3]=="(" and str(potential_counter) == str(file_name[-2]):#this means the
                        py_url = f'attachments/MC2 Data-{tab}_Attachment/{instance.replace("_", " ")}_Process Flow({potential_counter}){ending}'
                        if os.path.isfile(py_url):
                            # if  file_name.replace(instance_with_space,"") == f"({potential_counter})":
                            url = f'../../attachments/MC2 Data-{tab}_Attachment/{instance.replace("_", " ")}_Process Flow({potential_counter}){ending}'
                            a_img += [f'<img class="{imgclass}" src="{url}" >']

    return "".join(a_img)


def return_leftvsright():
    return """ <div class="">
      <div class="row">
      <div class="col-6" style="border-right: 1px solid #eee;">
          {}
        </div>
        <div class="col-6">
          {}
            </div>
        </div>
        </div>"""


def return_grid1():  # left vs right for poor people
    return """ <div class="mycard">
                  <div class="myleft">
                    {}
                  </div>
                  <div class="myright">
                    {}
                  </div>
                </div>
                """


def return_grid2(other_class="mb-5", card_class=''):
    return """ <div class="card """ + other_class + " " + card_class + """ ">
                  <div class="card-header">                    
                    {}
                  </div>
                  <div class="card-body """ + "" + """ ">
                    {}
                  </div>
                </div>
                """


def return_grid3():
    return """
                <div class="card" style="position: sticky; top: 0; z-index:99;">
                  <div class="card-header">
                    {}
                  </div>
                </div>
                """


def return_content_instance(instance, row, tab, dict_df):
    real_instance = instance

    cy1, component_cy_js = return_component_cy(dict_df=dict_df, highlight_classes=[tab],
                                               only_nodes=[x for x in dict_df.keys()],
                                               height="height50 width50 background2")
    label_direct_attribute = return_component_small_header("Table View")

    button1 = return_template_dropdown(id="filter_class", text="Class")
    button2 = return_template_dropdown(id="filter_instance", text="Instance")

    label_direct_compare = "<div class='row' >" + button1 + button2 + "</div><div id='versus'></div>"
    predefined_class = ",".join(return_string_indirect(tab))
    label_indirect_attribute = return_component_small_header("Graph View", id="predefined_relations",
                                                             h3_tag=f'data-predefined_relations="{predefined_class}"')
    label_indirect_attribute2 = return_component_small_header("Result")
    header = return_component_header(df=dict_df[tab], tab=tab, dict_df=dict_df, instance=real_instance)
    cy4 = return_cy4()
    if tab in ["Department"]:
        header = header + cy4

    spacer = return_component_spacer()
    template_card = return_template_card()
    cy2 = return_indirect_chart()
    cy2_text_result = get_cy2_result()
    layout_select = "<select class='' id='layoutselect'><option></option></select>"
    fullscreen_button = '<button class="btn btn-primary ml-3" id="fullscreen_button"> <i class="fa fa-expand"></i></button>'
    explainer = f'<ul><li><b>Left Click</b>: See indirect relations from {real_instance.replace("_", " ")}</li><li><b>Right Click</b>: Go to the Details Page of the selection.</li></ul>'
    modal = """
<div id="modal" class="modal modal-fullscreen-xl" id="modal-fullscreen-xl" tabindex="-1" role="dialog" aria-hidden="true">
  <div class="modal-dialog" role="document">
    <div class="modal-content">

      <div class="modal-body" id="modal_body">
        <div id='cy3' class=''></div>
      </div>
      <div class="modal-footer">
        <button id="modal_close" type="button" class="btn btn-primary" data-dismiss="modal" data-bs-dismiss="modal" >Close</button>  
      </div>
    </div>
  </div>
</div>
    """

    direct_part_left = (
                "<div class='width100' id='left_direct'>" + "" + "</div>")  # this function is now build in javascript
    direct_part_right = ("<div class='width100'>" + label_direct_compare + "</div>")

    grid1 = return_grid2(other_class="", card_class="background1")
    grid2 = return_grid2(other_class="mb-0", card_class="background2")
    leftrightgrid = return_leftvsright()

    process_image = return_instance_img(instance=instance, tab=tab, imgclass="instance_img")
    grid3label = return_component_small_header("Main Process", id="", h3_tag=f'')
    grid3 = return_grid2(other_class="mb-0", card_class="background2")

    a_subprocess_img = []
    # get subprocees pictures
    if tab in ["Employee_Process"]:
        a_subprocesses = dict_df[tab].at[instance, "Has Subprocess"]
        if isinstance(a_subprocesses, str):
            a_subprocesses = a_subprocesses.split(",")
            for subprocess in a_subprocesses:
                if pd.notna(subprocess):
                    # get picture
                    subprocess_image = return_instance_img(instance=subprocess, tab=tab, imgclass="instance_img")
                    a_subprocess_img += [
                        f"<h4><a href='{return_word_instance_url(class_tab=tab, instance=subprocess)}'> {len(a_subprocess_img) + 1}. Sub Process: {underscore_replacer(subprocess)}</a></h4>{subprocess_image}"]

    # create grid
    cysubprocess = f"<div id='cysubprocess' class=' mb-3 width100 background2' ></div>"
    cysubprocess =""
    combined_img = process_image+cysubprocess  # process has only one picture
    if len(a_subprocess_img) > 0:  # process has at least 1 picture in subprocess
        helper = [f"<div class='col col-md-4'>{x}</div>" for x in a_subprocess_img]
        grid_subprocess = '<div class="row mt-5">{}</div>'.format("".join(helper))
        combined_img += grid_subprocess


    # create all image combination
    if combined_img:
        grid3 = grid3.format(grid3label, "".join(combined_img))
    else:
        grid3 = ""



    # process_image=f'<img id="instance_img" src="{instance_img_url}" alt="{instance} Attached Image from Bitable">'+spacer if instance_img_url else ""
    direct_part = grid1.format(label_direct_attribute, leftrightgrid.format(direct_part_left, direct_part_right))
    indirect_part = grid2.format(label_indirect_attribute, leftrightgrid.format(explainer,
                                                                                layout_select + fullscreen_button + modal)) + f"<div class='background2'>{return_grid1().format(cy1, cy2)}</div>" + f"<div class='background2'>{return_grid1().format('', cy2_text_result)}</div>"

    content = header + spacer + grid3 + direct_part + spacer + indirect_part + spacer
    return template.format(content=content, jsinclude=component_cy_js)




def return_template_card():
    result = """
    <div class="card" style="width: 100%;">
      <div class="card-body">
        {}
      </div>
    </div>
    """
    return result


def return_component_spacer(default=0):
    return f'<p class="mt-{default}" style="margin-bottom:0.5rem;"></p>'


def get_cy2_result():
    return "<ul id='indirect_result' class='pb-5' ></ul>"


def return_indirect_chart(counter=2):
    return f"<div id='cy{counter}' class=' mb-3 height50 width50 background2'></div>"


def return_cy4():
    return f"<div id='cy4' class=' mb-3 height30 width100 background2'></div>"


def return_component_small_header(text="", component_inside='', id='', h3_tag=''):
    return f'<h3 id="{id}" {h3_tag} ><b>{text}{component_inside}</b></h3>'


def return_word_class_url(class_tab):
    return fr"../../page/{class_tab}/{class_tab}.html"


def return_word_instance_url(class_tab, instance):
    return fr"../../page/{class_tab}/{space_replacer(instance)}.html"


def return_component_header(df, tab, dict_df, instance, custom_header_text=""):
    try:
        count = len(df[df[tab].notna()])
        classcount = ""
    except:
        classcount = ""

    if instance:
        h1_icon = f'<a href="{return_word_class_url(class_tab=tab)}"><i class="fa-solid {return_string_icon(tab)} "></i></a>'
        h1_icon = f'<a class="text-secondary" href="{return_word_class_url(class_tab=tab)}"><i class="fa-solid {return_string_icon(tab)} text-secondary "></i><span class="underline"> {tab.replace("_", " ")}</span></a>'
        if tab=="Employee_Process":
            category=df.at[instance,'Process Category']
            if isinstance(category,str):
                categorylabel=category.replace('_', " ")
                category=f"<a class='text-secondary underline' href='{return_word_instance_url(class_tab='Capability',instance=category)}'>{categorylabel}</a>"
                category=f" • {category}"
            else:
                category=""
            modifiedby=df.at[instance,'Modified By']
            lastmodified=df.at[instance,'Last Modified']
            print(type(modifiedby))
            print(type(lastmodified))
            h1_icon=f"{h1_icon}<span class='text-secondary'>{category} • last modified by {modifiedby} on {lastmodified}</span>"
    else:
        h1_icon = f'<i class="fa-solid {return_string_icon(tab)} text-secondary "></i>'
    edit = f'<a href="{return_string_editurl(tab)}" style="font-size:1rem;" target="_blank" type="button" class="btn btn-primary btn-sm" ><i class="fa-solid fa-edit"></i></a>'
    other_classes = ""
    if instance:
        explainer = f''
    else:
        if tab not in ["User_Process", "Employee_Process"]:  # todo hardcoded
            explainer = f'<p class="text-secondary">{return_string_component(tab)}</p>'
        else:
            explainer = f''
    if not custom_header_text:
        header_text = f'{instance.replace("_", " ")}' if instance else "All " + tab.replace("_", " ")
    else:
        header_text = custom_header_text

    if instance:
        h1 = f'<h1 class="" id="header" data-current_class="{tab}"  data-current_instance="{instance}" >{maketitle(header_text)}  {edit} </h1> {h1_icon}' + explainer
    else:
        h1 = f'<h1 class="" id="header" data-current_class="{tab}"  data-current_instance="{instance}" >{h1_icon} {maketitle(header_text)} {edit} </h1>' + explainer

    grid3 = return_grid3()
    h1 = grid3.format(h1+" ")
    return return_component_spacer() + h1


def return_content_class(tab, df, dict_df):
    cards = ''
    for (fakekey, row), key in zip(df.iterrows(), df[tab]):
        if pd.isna(key) or key is None:
            continue

        cardstart = f'<div id="{space_replacer(key)}" class="card card_hover" style="width: 32%;float:left;">  <div class="card-body class_card_body"><h5 class="card-title"><a href ={f"../../page/{tab}/{key}.html"} >{key.replace("_", " ")}</a></h5>' + '{}</div></div>'
        cardmiddle = ""
        for row_key, row_item in row.items():
            if row_key != tab:
                row_entry = f"<li class='entry {space_replacer(row_key)}_entry'>{row_key}: {row_item}</li>"
                row_entry = ""
                cardmiddle = cardmiddle + row_entry

        cardmiddle = f"<ul class='card_ul' data-instance='{space_replacer(key)}_ul' id='{space_replacer(key)}_ul'>{cardmiddle}</ul>"
        cards = cards + cardstart.format(cardmiddle)

    # add filter button before card start
    component_filter, component_filter_js = return_component_filter(tab, df)
    component_filter = component_filter + "<hr/>"

    iframe = f"<iframe src='{return_string_gallery(tab)}' height='100%' width='100%'  style='margin-left=-50px !important; margin-right=-50px !important;'></iframe><hr/>"

    if "Department" in tab:
        iframe = "<iframe src='https://nio.feishu.cn/wiki/wikcnE50PKAKxW6u0IaIOyxyzTd#mindmap' height='100%' width='100%' ></iframe><hr/>" + iframe
    elif "Strategy" in tab:
        iframe = "<iframe src='https://nio.feishu.cn/wiki/wikcn5MZTGmEfOMR1HxJ45RrvVb#mindmap' height='100%' width='100%' ></iframe><hr/>" + iframe

    direct_relation_label = return_component_small_header(text=f"1. Specific {tab}:")
    indirect_relation_label = return_component_small_header(text="2. Relation to others:")
    header = return_component_header(df, tab, dict_df, instance="")
    template_card = return_template_card()
    # component_cy, component_cy_js=return_component_cy(dict_df=dict_df, highlight_classes=[tab], only_nodes=[tab]+return_array_related_classes(tab=tab,connections=2),height="height50")
    related_class = ""
    component_cy, component_cy_js = return_component_cy(dict_df=dict_df, highlight_classes=[tab],
                                                        only_nodes=related_class, height="height50")

    part_direct = (return_component_spacer() + iframe)
    spacer = return_component_spacer()
    grid2 = return_grid2()
    grid3 = return_grid3()
    part_direct = grid3.format(return_component_small_header("1. Class Overview"), "") + component_cy
    part_direct = ""
    part_inddirect = grid2.format(return_component_small_header("1. Instance Overview"), cards)

    content = header + spacer + part_direct + spacer + part_inddirect + spacer
    return template.format(content=content,
                           jsinclude=component_cy_js + component_filter_js)


def return_content_index(dict_df):
    h1 = f'<h2 class="pt-5 pb-1" id="header" data-current_class="index"><i class="fa-solid fa-face-smile-wink fa-xl"></i> What is to mc² system?</h2>'
    explainer = f'<p class=" pb-3 text-secondary">mc² is a system made by <b>CJ</b> to understand complex relations within NIO Europe: such as process relations, user journey, strategy to process relations and more. Mc² stands for Energy and is derived from the famous equation E=mc². The mission is to enable all people and give them energy.</p><hr/>'
    p = f'<p class=" text-secondary">  <ul><b>How to Use:</b><li> <b>Left Click</b>: move entity around.</li><li><b>Right Click</b>: jump to the details page.</li><li><b>Mouse Wheel</b>: zoom in and out.</li></ul></p>'
    p2 = f'<p class=" text-secondary">  <ul><b>Useful Examples:</b><li> <b>Employee -> Employee Process -> KPI</b>: Show all relevant KPIs to one employee from his perspective. </li><li><b>KPI -> Employee Process -> Employee</b>: See all relevant People related to one KPI. Useful for leaders. </li><li><b>Role -> Employee -> System</b>: See all systems that a particular role is using</li></ul></p>'

    cy, chart_js = return_component_cy(dict_df, highlight_classes=[x for x in dict_df.keys()],
                                       only_nodes=[x for x in dict_df.keys()])
    return template.format(content=h1 + explainer + p + cy + p2, jsinclude=chart_js)


def get_user_journey_banner(user_journey_instance):
    if not user_journey_instance:
        return ""
    for file_path in glob.glob("img/*.jpg"):
        file_name = os.path.basename(file_path)
        file_name_withoutjpg = os.path.basename(file_name).replace(".jpg", "")
        if file_name_withoutjpg in user_journey_instance:
            if ".jpg" in file_name:
                return f"../../img/{file_name}"
    return ""


dict_link = {}


# calculate then linkable column once, then reuse it again in python and javascript
def generate_linkabe_column():
    df_result = pd.DataFrame()
    # dict_df,wb=get_raw_dict_df()#get raw dict df, because dict df because cleanup replaces the string and data
    dict_df, wb = get_dict_df()  # get raw dict df, because dict df because cleanup replaces the string and data
    confidence_level = 0.6
    dict_result = {}

    for tab, df in dict_df.items():  # source tab
        for tab_compare in dict_df.keys():  # check against destination tab
            for col in df.columns:  # for each column of the source tab
                result = 0
                itemchecked = 0
                DataSeries = dict_df[tab_compare][tab_compare]
                for cell_array in df[col]:
                    if not isinstance(cell_array, str):  # continue if we data is float
                        continue

                    if not cell_array:
                        continue

                    if pd.isna(cell_array):
                        continue

                    for cell in cell_array.split(","):  # todo split by comma could be messed up with uncleaned data
                        if cell in DataSeries.values:
                            result = result + 1
                            itemchecked = itemchecked + 1
                        else:
                            result = result + 0
                            itemchecked = itemchecked + 1

                if itemchecked != 0:
                    df_result.at[tab + "_" + str(col), tab_compare] = result / itemchecked

                    # hard coded edges to exclude
                    if col in ["Has Leader", "Is Leader of Department"]:  # todo hard coded label here
                        continue

                    if result / itemchecked >= confidence_level and tab_compare != col:
                        if not tab in dict_result:
                            dict_result[tab] = {col: tab_compare}
                        else:
                            last_dict = dict_result[tab]
                            last_dict[col] = tab_compare
                            dict_result[tab] = last_dict  # store it somewhere


    #hardcoded to remove links
    df_result.at["Department_Capability","Capability"]=0
    df_result.at["Department_Belongs_to_Topic","Topic"]=0
    df_result.at["Capability_Defined_by_Department","Department"]=0
    df_result.at["Capability_Belongs_to_Topic","Topic"]=0

    df_result.to_excel("wlink.xlsx")#store as excel

    # save as global variable
    global dict_link
    dict_link = dict_result

    # store it as json
    json_result = json.dumps(dict_result, indent=4)
    global link
    link = dict_result
    with open(fr"bootstrap/js/link.js", "w", encoding="utf-8") as file:
        file.write(f"var link = {json_result}")
    return df_result


# for index_class, check if all his columns are linkable
def get_linkable_cols(instance_class, return_type=1):
    helper = link[instance_class]
    if return_type == 1:  # return everything {"my naming": "real label", "my naming": "real label"}
        return helper
    elif return_type == 2:  # only return dict keys
        return helper.keys()
    elif return_type == 3:  # only return values
        return helper.values()
    else:
        return helper


def col_linked(instance_class, col, type=2):
    all_links_of_class = get_linkable_cols(instance_class=instance_class, return_type=1)
    if col in all_links_of_class.keys():
        if type == 1:  # return bool two times 1 is correct
            return True
        else:  # return label
            return all_links_of_class[col]
    else:
        if type == 1:  # return bool two times 1 is correct
            return False
        else:  # return label
            return ""


def return_content_user_journey(dict_df, tab="User_Journey", one_bm="Subscription"):
    note = """
            <main>
  <nav class="section-nav">
		<ol>
			<li><a href="#introduction">Introduction</a></li>
			<li><a href="#request-response">Request &amp; Response</a></li>
			<li><a href="#authentication">Authentication</a></li>
			<li><a href="#endpoints">Endpoints</a>
				<ul>
					<li class=""><a href="#endpoints--root">Root</a></li>
					<li class=""><a href="#endpoints--cities-overview">Cities Overview</a></li>
					<li class=""><a href="#endpoints--city-detail">City Detail</a></li>
					<li class=""><a href="#endpoints--city-config">City Config</a></li>
					<li class=""><a href="#endpoints--city-spots-overview">City Spots Overview</a></li>
					<li class=""><a href="#endpoints--city-spot-detail">City Spot Detail</a></li>
					<li class=""><a href="#endpoints--city-icons-overview">City Icons Overview</a></li>
					<li class=""><a href="#endpoints--city-icon-detail">City Icon Detail</a></li>
				</ul>
			</li>
			<li class=""><a href="#links">Links</a></li>
			<li class=""><a href="#expanders">Expanders</a></li>
			<li class=""><a href="#filters">Filters</a></li>
		</ol>
	</nav>


	<div>
		<h1>User Journey</h1>
		<section id="introduction">
			<h2>Introduction</h2>
			<p>…</p>
		</section>
		<section id="request-response">
			<h2>Request &amp; Response</h2>
			<p>…</p>
		</section>
		<section id="authentication">
			<h2>Authentication</h2>
			<p>…</p>
		</section>
		<section id="endpoints">
			<h2>Endpoints</h2>
			<section id="endpoints--root">
				<h3>Root</h3>
				<p>…</p>
			</section>
			<section id="endpoints--cities-overview">
				<h3>Cities Overview</h3>
				<p>…</p>
			</section>
			<section id="endpoints--city-detail">
				<h3>City Detail</h3>
				<p>…</p>
			</section>
			<section id="endpoints--city-config">
				<h3>City Config</h3>
				<p>…</p>
			</section>
			<section id="endpoints--city-spots-overview">
				<h3>City Spots Overview</h3>
				<p>…</p>
			</section>
			<section id="endpoints--city-spot-detail">
				<h3>City Spot Detail</h3>
				<p>…</p>
			</section>
			<section id="endpoints--city-icons-overview">
				<h3>City Icons Overview</h3>
				<p>…</p>
			</section>
			<section id="endpoints--city-icon-detail">
				<h3>City Icon Detail</h3>
				<p>…</p>
			</section>
		</section>
		<section id="links">
			<h2>Links</h2>
			<p>…</p>
		</section>
		<section id="expanders">
			<h2>Expanders</h2>
			<p>…</p>
		</section>
		<section id="filters">
			<h2>Filters</h2>
			<p>…</p>
		</section>
	</div>


</main>

    """

    user_journey_template = """
    <main id='main' data-forjourney="{}" data-forperspective="{}">
      <nav class="section-nav" style="margin-right:20px;">
            <ol > {}</ol>
      </nav>
    	<div class="p-0" id="speechify_content" >
    		""" + return_component_header(df=dict_df[tab], tab=tab, dict_df="", instance="",
                                          custom_header_text=f"{one_bm} {tab.replace('_', ' ')}") + """
    		<p class="mb-5" ><p>
    		{}
    	</div>
    </main>
        """

    # now only checking user journey
    df_user_or_employee_journey = dict_df[tab]
    if tab == "User_Journey":
        process_display_name = "For User_Process"  # my display name Todo hard coded, to replace with custom label
        process_display_name = "Has User Processes"  # my display name Todo hard coded, to replace with custom label
        process_id_name = col_linked(instance_class=tab, col=process_display_name)  # real id
    else:
        process_display_name = "Has Employee Processes"  # my display name
        process_id_name = col_linked(instance_class=tab, col=process_display_name)  # real id

    all_nav = ""
    all_section = ""
    for journey_counter, (key, row) in enumerate(df_user_or_employee_journey.iterrows()):
        journey_counter = 0  # quick and dirty way to reset it for h2 labeling
        if pd.isna(key):
            continue

        if one_bm in key:
            a_user_process = df_user_or_employee_journey.at[key, process_display_name]

            # if this user journey has 1 or many processes, then the navvalue and secitonvalue will be replaced
            # if user journey has 0 process. then it stays as key

            if isinstance(a_user_process, str):
                if len(a_user_process.split(",")) > 0:
                    navvalue_content = ""
                    # calculate
                    for process_counter, process in enumerate(a_user_process.split(",")):
                        process_display = process.replace(" ", "_")
                        navvalue_content = navvalue_content + f'<li class=""><a href="#{process_display}">{process_display.replace("_", " ")}</a></li>'
                    navvalue = f"<ul>{navvalue_content}</ul>"

                    sectionvalue = ""

                    for process_counter, process in enumerate(a_user_process.split(",")):
                        label = f"{journey_counter + 1}.{process_counter + 1} "
                        process_display = process.replace(" ", "_")
                        link = f"<a href='../../page/{col_linked(instance_class=tab, col=process_display_name)}/{process.replace(' ', '_')}.html'>{label + process.replace('_', ' ')}</a>"
                        image = return_instance_img(instance=underscore_replacer(process), tab=process_id_name,
                                                    imgclass="journeyimg")
                        # image=f'<img class="journeyimg" src="../../attachments/MC2 Data-{process_name_without}_Attachment/{process.replace("_"," ")}_Process Flow.jpg" >'

                        # get summary of the process

                        try:
                            df_process = dict_df[process_id_name]
                            summary = df_process.at[process, "Process Summary"]
                        except:
                            print(journey_counter, key, process_counter, process)
                            summary = ""
                        sectionvalue = sectionvalue + f"<section id='{process_display}'><h2>{link}</h2>{image}{summary}</section>"

                else:
                    navvalue = ""
                    sectionvalue = key
            else:
                navvalue = ""
                sectionvalue = key

            navid = key
            sectionid = key
            # for each user journey, get their user process
            link = f"<a href='../../page/{tab}/{sectionid.replace(' ', '_')}.html'>{journey_counter + 1}. {sectionid.replace('_', ' ')}</a>"
            nav_element = f"<li><a href='#{navid}'>{navid.replace('_', ' ')}</a> {navvalue}</li>"
            user_journey_banner = f'<img class="journeyimg" src="{get_user_journey_banner(key)}" >' + return_component_spacer(
                default=5)
            section_element = f"<section id='{sectionid}'><h2>{link}</h2><hr/>{user_journey_banner}{sectionvalue}</section>"
            all_nav = all_nav + nav_element
            all_section = all_section + section_element

    content = user_journey_template.format(one_bm, tab, all_nav, all_section)
    return template.format(content=content, jsinclude="")


# not used atm
def return_component_filter(tab, df):
    template_filter = """<div class="dropdown"> <button class="btn btn-primary dropdown-toggle" type="button" id="dropdownMenuButton" style="width:100%;"    data-bs-toggle="dropdown" aria-expanded="false">Select Direct Attribute</button>   <ul class="dropdown-menu" id='filter_ul' aria-labelledby="dropdownMenuButton">       {}    </ul>    </div>"""
    items = ""
    for col in df.columns:
        if col != tab:
            item = f"""<li class="dropdown-item"><div class="form-check">
                    <input class="form-check-input big-checkbox" type="checkbox" value="{space_replacer(col)}_filter" id="{space_replacer(col)}_filter" checked />
                    <label class="form-check-label" for="{space_replacer(col)}_filter">&nbsp;{col}</label>
                </div></li>"""
            items = items + item

    js_part = ""
    for col in df.columns:
        if col != tab:
            js_part_entry = """ $('#""" + f"{space_replacer(col)}_filter" + """').change(function() {if(this.checked) { $('.""" + f"{space_replacer(col)}_entry" + """').show();}else{  $(".""" + f"{space_replacer(col)}_entry" + """").hide();  } }); """
            js_part = js_part + js_part_entry
    js_part = f"<script>{js_part}</script>"

    indirect_class = """
    <select class="" aria-label="Default select example" id="indirect_class">
    <option selected>Select Indirect Attribute</option>
    </select>
    """

    indirect_attribute = """
        <select class="" aria-label="Default select example" id="indirect_attribute">
        </select>
        """

    return [template_filter.format(items), js_part]


def return_template_dropdown(id, text):
    return """<select class="" id='""" + id + """'></select>"""


def return_component_cy(dict_df, only_nodes=[], highlight_classes=["Employee"], height="height100",
                        add_instance_label=""):
    """cy=cytoscape.js"""
    cy = f"<div id='cy' class=' mb-3 {height}'></div>"
    return [cy, f""]


def reverse_dict(my_dict):
    inverted_dict = {value: key for key, value in my_dict.items()}
    return inverted_dict


def output_file(content, filename):
    with open(fr"{filename}", "w", encoding="utf-8") as file:
        file.write(str(content))


def create_html():
    # init
    dict_df, wb = get_dict_df()

    # calculate progress tracking
    # by By department Who has finished how many
    """
    Do two visualizations

    1. loop over all process
    2. filter out things that I created:
    3.1 group by [Capbility, L2 Department, L3 Department ]
    3.2 group by [topic ,  capability]

    """

    def get_column_name(tab, col_name):
        global dict_link
        d_helper = dict_link[tab]
        for key, val in d_helper.items():
            if val == col_name:
                return key
        else:
            return ""

    def grouped(iterable, n):
        "s -> (s0,s1,s2,...sn-1), (sn,sn+1,sn+2,...s2n-1), (s2n,s2n+1,s2n+2,...s3n-1), ..."
        return zip(*[iter(iterable)] * n)

    def pairwise(lst):
        "[(1, 2), (2, 3), (3, 4), (4, 5), (5, 6)] and not [(1, 2),  (3, 4),  (5, 6)]"
        return [(lst[i], lst[i + 1]) for i, _ in enumerate(lst[:-1])]

    def get_departments(dict_df, topic):
        df_topic = dict_df["Topic"]
        print(df_topic)
        df_department = df_topic[df_topic["Topic"] == topic]
        print(df_department)
        return df_department["Defined by Department"].unique()


    def get_link_people(people):
        urlt = "http://mc2.nioint.com/page/Employee/{}.html#Employee_Process"
        return  f"<a target='_blank' href='{urlt.format(people.lower().replace(' ','_'))}'>@{people}</a>"



    if False:
        # generate communication email
        """
        1. join all tables together (add capability, Department)
        2. Group by whatever you want
        """
        df_merged = pd.DataFrame()
        for tab_left, tab_right in pairwise(["Employee_Process", "Capability", "Department"]):
            if df_merged.empty:
                df_left = dict_df[tab_left]
            else:
                df_left = df_merged
            df_right = dict_df[tab_right]
            tab_name_left = get_column_name(tab_left, tab_right)

            # rename column to add tab in advance
            df_merged = pd.merge(left=df_left, right=df_right, how='left', left_on=f"{tab_name_left}",                            right_on=f"{tab_right}",                             left_index=False, suffixes=(f'_{tab_left}', f'_{tab_right}'))

        #replace underscore
        for col in df_merged:
            try:
                df_merged[col] = df_merged[col].str.replace("_", " ")
            except:
                pass
        df_merged.to_excel(f"z merged.xlsx") #df_merged finished

        # generate report
        a_g = ['Belongs to Topic', 'Defined by Department', "Has Leader"]
        email = ""
        template_topic = get("template/topic.html")
        template_dpt = get("template/dpt.html")


        # groupby topic with  3
        for counter in [1,2,3]:
            a_g = ['Belongs to Topic', 'Defined by Department', "Has Leader"]
            a_g=a_g[0:counter]
            df_grouped_topic = df_merged.groupby(a_g).count()
            df_grouped_topic = df_grouped_topic.reset_index()
            df_grouped_topic=df_grouped_topic.sort_values(a_g[0],ascending=False)
            df_grouped_topic.to_excel(f"z topic{counter}.xlsx")

        a_g = ['Belongs to Topic', 'Defined by Department', "Has Leader"]



    # looping over all topics
    if False:
        for counter, atopic in enumerate(df_grouped_topic[a_g[0]].unique()):
            print(f"Department {atopic}")
            # for each unique group1element find all
            df_group1_f=df_grouped_topic[df_grouped_topic[a_g[0]]==atopic]

            # looping over topic
            a_dpt_result=[]
            for index, topic,dpt,leader in zip(df_group1_f.index, df_group1_f[a_g[0]], df_group1_f[a_g[1]], df_group1_f["Has Leader"]):
                leader=leader.title()

                de_l=df_merged.loc[df_merged[a_g[1]]==dpt,"DE Counterpart"].iat[0]
                dk_l=df_merged.loc[df_merged[a_g[1]]==dpt,"DK Counterpart"].iat[0]
                nl_l=df_merged.loc[df_merged[a_g[1]]==dpt,"NL Counterpart"].iat[0]
                se_l=df_merged.loc[df_merged[a_g[1]]==dpt,"SE Counterpart"].iat[0]
                no_l=df_merged.loc[df_merged[a_g[1]]==dpt,"NO Counterpart"].iat[0]

                de_c=0
                dk_c=0
                nl_c=0
                se_c=0
                no_c=0

                de_h,dk_h,nl_h,se_h,no_h =[get_link_people(x) for x in [de_l,dk_l,nl_l,se_l,no_l]]

                de=f"<strong>{de_c}</strong> {de_h}"
                dk=f"<strong>{dk_c}</strong> {dk_h}"
                nl=f"<strong>{nl_c}</strong> {nl_h}"
                se=f"<strong>{se_c}</strong> {se_h}"
                no=f"<strong>{no_c}</strong> {no_h}"

                listed = len(df_merged[(df_merged[a_g[0]] == topic) & (df_merged[a_g[1]] == dpt)])
                listed = f"<strong>{listed}</strong> {get_link_people(leader)}"
                leadersigned=0
                leadersigned = f"<strong>{leadersigned}</strong> {get_link_people(leader)}"

                #put everything together
                href = f"http://mc2.nioint.com/page/Department/{dpt.replace(' ','_')}.html#Employee_Process"
                dpt_info = template_dpt.format(dpt=dpt,  listed=listed, signed1=leadersigned, href=href, de=de,no=no,nl=nl,dk=dk,se=se)
                a_dpt_result += [dpt_info]


            content = f"<ul>{''.join(a_dpt_result)}</ul>"

            email += template_topic.format(topic=f"{counter+1}. {atopic}", content=content)
        output_file(email, "email.html")


    # prevent duplicates of process names and duplicates of other tabs
    for tab, df in dict_df.items():
        xs = list(df[tab])
        result = set([x for x in xs if xs.count(x) > 1])
        if result:
            print(f"DUPLICATES of {tab}: {result}")

    # classes
    dict_extrawurst = {}
    for tab, df in dict_df.items():
        Path(f"page/{tab}").mkdir(parents=True, exist_ok=True)
        if tab in dict_extrawurst:
            result = dict_extrawurst[tab](tab, df, dict_df)
        else:
            result = return_content_class(tab=tab, df=df, dict_df=dict_df)
        with open(fr"page/{tab}/{tab}.html", "w", encoding="utf-8") as file:
            file.write(str(result))

    # instances
    for tab, df in dict_df.items():
        Path(f"page/{tab}").mkdir(parents=True, exist_ok=True)
        try:
            df.set_index(tab, inplace=True)
        except:
            pass
        for instance, row in df.iterrows():
            if instance and not pd.isna(instance) and instance is not None:
                result = return_content_instance(instance=instance, row=row, tab=tab, dict_df=dict_df)
                # change the word primary to success
                result = result.replace("primary", 'primary')
                try:
                    with open(fr"page/{tab}/{instance}.html", "w", encoding="utf-8") as file:
                        file.write(str(result))
                except:
                    print(f"ERROR saving instance page {tab}: {instance}")

    # index html
    Path(f"page/index").mkdir(parents=True, exist_ok=True)
    result = return_content_index(dict_df)
    with open(fr"page/index/index.html", "w", encoding="utf-8") as file:
        file.write(str(result))

    # user journey specific Page, this comes after the general creation and overwrites the class html pages
    # for user_employee_journey in ["User_Journey", "Process_Category"]:
    for user_employee_journey in []:  # no user journey override
        a_bm = ["Purchase", "Subscription", "Op-Leasing", "Fin-Leasing"]
        for one_bm in a_bm:
            Path(f"page/{user_employee_journey}").mkdir(parents=True, exist_ok=True)
            result = return_content_user_journey(dict_df, tab=user_employee_journey, one_bm=one_bm)
            with open(fr"page/{user_employee_journey}/{user_employee_journey}_{one_bm}.html", "w",
                      encoding="utf-8") as file:
                file.write(str(result))

            # quit and dirt solution to store it also as general user journey so that other pages can access this page
            with open(fr"page/{user_employee_journey}/{user_employee_journey}.html", "w", encoding="utf-8") as file:
                file.write(str(result))


logged_in = False


def get_all_wikis(id, driver):
    """
    takes a url and returns a set of wiki id
    :param url:
    :return:
    """
    url = f"https://nio.feishu.cn/wiki/{id}"
    driver.get(url)
    a_results = []

    global logged_in
    if not logged_in:
        input("Enter after logged in")
        logged_in = True

    soup = BeautifulSoup(driver.page_source)
    for tag in soup.find_all(class_='tree-item'):
        # for tag in driver.find_all("css selector", ".tree-item"):
        # get the direct div
        # div1=tag.select_one('.tree-node-wrapper.wiki-tree-normal-node')
        # div2=div1.select_one('.tree-item')
        div2 = tag

        token = div2.attrs["data-token"]
        title = div2.attrs["data-title"]
        parent = div2.attrs["data-parent"]
        pos = div2.attrs["data-pos"]
        a_results += [token]

    return a_results


def get(path="h1.html", way=3):
    if way == 1:  # get path for python
        return f"{path}"
    elif way == 2:  # get path for html rendering like img path
        return f"../{path}"
    elif way == 3:  # get file content itself
        return open(f"{path}", "r", encoding='utf-8').read()


def initiate_driver():
    driver = uc.Chrome(version_main=108)
    driver.maximize_window()  # maximize window size
    return driver


def rekursive_crawl_wiki(start_id="", driver="", found_result=[]):
    """
    crawl through wiki to get all articles
    1. loop over all articles found in the current page
    2. check if these articles are already in the found_result
        2.1 if article is found first time, then also rekursive crawl that wiki

    1. termination condition is when the page has no more new urls

    use ID, system calculates url automatically. more readable than url
    :return:
    """
    if not driver:
        driver = initiate_driver()

    for found_id in get_all_wikis(id=start_id, driver=driver):
        """a found url can be already known, new and not child node, new and child node"""
        if found_id not in found_result:  # if found, the the wik must be child of page article
            # found_result=found_result.union(get_all_wikis(found_url), found_result=found_result)
            found_result += [found_id]
            found_result += [rekursive_crawl_wiki(start_id=found_id, driver=driver, found_result=found_result)]

    return found_result


def craw_wiki():
    # find all the token from EU wiki, or other wiki
    eu_wiki = "wikcnC8uxuVJ2C12E7lTdDUvXKe"
    threec_wki = "wikcnibclnFGSyFOvzn9nVtosre"
    a_tokens = rekursive_crawl_wiki(start_id=threec_wki)

    # analyse them and create keyword for that article.

    # in creating the detailed article,check if there is any high correlation


if __name__ == '__main__':
    # craw_wiki()
    # 明显错的

    if True:
        generate_linkabe_column()
        print("FINISHED generate_linkabe_column")
        generate_lark_to_mc2_link()
        print("FINISHED generate_lark_to_mc2_link")
        return_global_navbar()
        return_global_html()
        create_html()





