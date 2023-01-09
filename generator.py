import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
import selenium
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
import os.path
from selenium import webdriver
import time
import glob, os
import os
import openpyxl
import re
#import pinyin
from pypinyin import pinyin, lazy_pinyin, Style
import shutil
from PIL import Image

from pathlib import Path
from pandas.api.types import is_string_dtype
from bs4 import BeautifulSoup
from pandas.api.types import is_numeric_dtype
from deep_translator import (GoogleTranslator,
                             MicrosoftTranslator,
                             PonsTranslator,
                             LingueeTranslator,
                             MyMemoryTranslator,
                             YandexTranslator,
                             PapagoTranslator,
                             DeeplTranslator,
                             QcriTranslator,
                             single_detection,
                             batch_detection)



"""
this py generates visualized html chart based on table input from lark

"""

def by_Strategy():
   """
   read the data.xlsx that is stored in the same path as the py file
   go through all tabs to check and select capability
   depth first or wide first
   use lark to visualize them. Friking awesome

   Currently only supporting 1:1 relation
   Future: 1:N Relation, N:N Relation

   Order:
   Strategy
   Capability
   Process
   Department
   People

   :return:
   """

   for excel_data_raw in glob.glob("*.xlsx"):
       dict_df = pd.read_excel(excel_data_raw, sheet_name=None)
       break

   soup = BeautifulSoup("""<html></head><body><ul></ul></body></html>""", "html.parser")

    # initial leaf
   for name,df_sheet in dict_df.items():
       if name in ['Strategy']:
           for item_name in df_sheet[df_sheet.columns[0]]:
               if item_name is not None and not pd.isna(item_name):
                   print(f"strategy {item_name}")
                   li = soup.new_tag("li")
                   li.string=f"{name}: {item_name}"
                   li.attrs["id"]=item_name
                   soup.html.body.ul.insert_after(li)

   print(" ")
   print(" ")
   dict_looper={
       "Capability":"For Strategy",
       "Process":"For Capability",
       "Department":"For Process",
   }




   for key,val in dict_looper.items():
       for name, df_sheet in dict_df.items():
           if name in [key]:

               #load openpyxl to extract links from file
               wb = openpyxl.load_workbook(excel_data_raw)
               ws = wb.get_sheet_by_name(key)

               for row, (item_name, for_layer) in enumerate(zip(df_sheet[df_sheet.columns[0]], df_sheet[f"{val}"])):
                   try:
                       hyperlink=ws.cell(row=row+2, column=1).hyperlink.target
                       if "feishu" in hyperlink:
                           pass
                       else:
                           hyperlink = ""
                   except:
                       hyperlink=""

                   if not pd. isna(item_name) and not pd.isna(for_layer):
                       print(f"{item_name}:  for {for_layer}")
                       #check if this element has a url, if not, create one

                       ul_strategy=soup.find("ul", {"id": f"ul_{for_layer}"})
                       if ul_strategy is None:
                           ul_strategy = soup.new_tag("ul")
                           ul_strategy.attrs["id"] = f"ul_{for_layer}"
                           insert_li = soup.find("li", {"id": for_layer})
                           insert_li.insert(1, ul_strategy)

                       #create li and insert in ul inside the li
                       if hyperlink:
                           a = soup.new_tag("a")
                           a.string = f"{key}: {item_name}"
                           a.attrs["href"] = hyperlink
                           a.attrs["id"] = item_name
                           li = soup.new_tag("li")
                           li.insert(0, a)
                           ul_strategy.insert(0, li)
                       else:
                           li = soup.new_tag("li")
                           li.string = f"{key}: {item_name} {hyperlink}"
                           li.attrs["id"]=item_name
                           ul_strategy.insert(0,li)
       print(" ")
       print(" ")


   with open("by_Strategy.html", "w",encoding="utf-8") as file:
       file.write(str(soup))





def by_Department():
   """


   :return:
   """

   for excel_data_raw in glob.glob("*.xlsx"):
       dict_df = pd.read_excel(excel_data_raw, sheet_name=None)
       break

   soup = BeautifulSoup("""<html></head><body><ul></ul></body></html>""", "html.parser")

    # L2 Department
   for name,df_sheet in dict_df.items():
       if name in ['L2 Department']:
           for item_name in df_sheet[df_sheet.columns[0]]:
               if item_name is not None and not pd.isna(item_name):
                   print(f"L2 Department {item_name}")
                   li = soup.new_tag("li")
                   li.string=f"{name}: {item_name}"
                   li.attrs["id"]=item_name
                   soup.html.body.ul.insert_after(li)

   print(" ")
   print(" ")

   """
   L2
   L3
   Process map
   """

   dict_looper={
       "L3 Department":"For L2 Department",
   }
   for key,val in dict_looper.items():
       for name, df_sheet in dict_df.items():
           if name in [key]:

               #load openpyxl to extract links from file
               wb = openpyxl.load_workbook(excel_data_raw)
               ws = wb.get_sheet_by_name(key)

               for row, (item_name, for_layer) in enumerate(zip(df_sheet[df_sheet.columns[0]], df_sheet[f"{val}"])):
                   try:
                       hyperlink=ws.cell(row=row+2, column=1).hyperlink.target
                       if "feishu" in hyperlink:
                           pass
                       else:
                           hyperlink = ""
                   except:
                       hyperlink=""

                   #add leader
                   leader = df_sheet.at[row, "Leader"]



                   if not pd. isna(item_name) and not pd.isna(for_layer):
                       print(f"{item_name}:  for {for_layer}")
                       #check if this element has a url, if not, create one

                       ul_strategy=soup.find("ul", {"id": f"ul_{for_layer}"})
                       if ul_strategy is None:
                           ul_strategy = soup.new_tag("ul")
                           ul_strategy.attrs["id"] = f"ul_{for_layer}"
                           insert_li = soup.find("li", {"id": for_layer})
                           insert_li.insert(1, ul_strategy)

                       #create li and insert in ul inside the li
                       if hyperlink:
                           a = soup.new_tag("a")
                           a.string = f"{key}: {item_name} by {leader}"
                           a.attrs["href"] = hyperlink
                           a.attrs["id"] = item_name
                           li = soup.new_tag("li")
                           li.insert(0, a)
                           ul_strategy.insert(0, li)
                       else:
                           li = soup.new_tag("li")
                           li.string = f"{key}: {item_name} ({leader})" if not pd.isna(leader) else f"{key}: {item_name}"
                           li.attrs["id"]=item_name
                           ul_strategy.insert(0,li)



       # Hard code process
       wb = openpyxl.load_workbook(excel_data_raw)
       ws = wb.get_sheet_by_name("Process")
       # go through all process and attach to L3 Departments
       # add process for each L3 Department, iterate through all process and check if there is l3 in it
       for process_name, df_process in dict_df.items():
           if process_name in ["Process"]:
               for row, (index, process, L3_string) in enumerate(zip(df_process.index, df_process["Process"], df_process["By L3 Department"])):
                   # check if comam is in string TODO add consolidated departments process
                   try:
                       hyperlink = ws.cell(row=row + 2, column=1).hyperlink.target
                       if "feishu" in hyperlink:
                           pass
                       else:
                           hyperlink = ""
                   except:
                       hyperlink = ""


                   if not pd.isna(process) and not pd.isna(L3_string):
                       ul_L3 = soup.find("ul", {"id": f"ul_{L3_string}"})
                       if ul_L3 is None:
                           ul_L3 = soup.new_tag("ul")
                           ul_L3.attrs["id"] = f"ul_{L3_string}"
                           insert_li = soup.find("li", {"id": L3_string})
                           try:
                               insert_li.insert(1, ul_L3)
                           except:
                               continue
                        # insert li into UL by equals adding process to department

                       if hyperlink:
                           a = soup.new_tag("a")
                           a.string = f"Process: {process}"
                           a.attrs["href"] = hyperlink
                           a.attrs["id"] = process
                           li = soup.new_tag("li")
                           li.insert(0, a)
                           ul_L3.insert(0, li)
                       else:
                           li = soup.new_tag("li")
                           li.string = f"Process: {process}"
                           li.attrs["id"]=process
                           ul_L3.insert(0,li)


       print(" ")
       print(" ")


   with open("by_Department.html", "w",encoding="utf-8") as file:
       file.write(str(soup))


def User_Process():
    """
    1. create html file content using python (xlsx -> python -> html -> github)
    2. paste it to index.html template
    3. upload it to github
    =half automation vizualization
    """

    for excel_data_raw in glob.glob("*.xlsx"):
        dict_df = pd.read_excel(excel_data_raw, sheet_name=None)
        break

    # Hard code process
    wb = openpyxl.load_workbook(excel_data_raw)
    ws = wb.get_sheet_by_name("User Process")

   #do first row template
    part_one=f'<div class="m30"><div class="row"><div class="col-6" id=one><h3>Front End</h3><div class="row">'
    middle=f''
    part_two=f'</div></div> <div class="col-6" id=two><h3>Back End</h3><div class="row"></div></div> </div></div> '



    # for each user journey create one row
    df=dict_df['User Process']
    for category in df["Category"].unique():
        try:
            #if int(category[0])>0:
            if True:
                mitte=[]
                # select all tasks that this cateogry has
                for row, (user_process, category_filter) in enumerate(zip( df["User Process"], df.loc[df["Category"]==category,"Category"])):
                    try:
                        hyperlink = ws.cell(row=row + 2, column=1).hyperlink.target
                        if "nio" in hyperlink and "feishu" in hyperlink:
                            mitte += [f"<li>{user_process}:  <a href='{hyperlink}' target='_blank'>as User</a> , <a href='{hyperlink}' target='_blank'>as NIO</a></li>"]
                        else:
                            mitte += [f"<li>{user_process}:  <a> as User</a></li>"]
                    except:
                        mitte += [f"<li>{user_process}:  <a > as User</a></li>"]


                #generate ul list under each body content

                mitte="<ul>"+"".join(mitte)+"</ul>"

                if category[0]=="5":
                    col="4"
                else:
                    col="12"
                opener = f'<div class="col-1 userjourneycat">'
                opener=""
                h2 = f'<div class="accordion-header" id="h2_{category}"><button class="accordion-button" type="button" data-bs-toggle="collapse" data-bs-target="#collapse_{category}" aria-expanded="true" aria-controls="collapse_{category}">{category}</button></div>'
                according_body = f'<div id="collapse_{category}" class="accordion-collapse collapse" aria-labelledby="h2_{category}" data-bs-parent="#{category}"><div class="accordion-body">{mitte}</div></div>'
                according_div = f'<div class="accordion col-{col} userjourneycat content_font" id="{category}"> <div class="accordion-item">{h2}{according_body}'
                end = '</div></div>'
                result=opener+according_div+end
                middle = middle + result
        except:
            pass

    print("middle: ",middle)

    row1=part_one+middle+part_two


    #row1='<div class="container"> <div class="row"> <div class="col-sm"> One of three columns </div> <div class="col-sm"> One of three columns </div> <div class="col-sm"> One of three columns </div> </div></div>'






    other_row='<div class="container"> <div class="row"> <div class="col-sm"> One of three columns </div> <div class="col-sm"> One of three columns </div> <div class="col-sm"> One of three columns </div> </div></div>'
    other_row =""
    result=f'{row1}{other_row}'

    return result


def strategy_purchase():
    return '<iframe width="100%" height="600" src="https://nio.feishu.cn/wiki/wikcn5MZTGmEfOMR1HxJ45RrvVb#mindmap" name="iframe_a" title="Iframe Example"></iframe>'


def department_purchase():
    return '<iframe width="100%" height="600" src="https://nio.feishu.cn/wiki/wikcnE50PKAKxW6u0IaIOyxyzTd" name="iframe_a" title="Iframe Example"></iframe>'


def content_generator():
    """generates div that gets pasted as content"""
    perspective = ["Private_User", "SME_User", "Big_Corp_User", "Employee", "User_And_Employee", "Department",
                   "Strategy", "Process"]
    ownership=["Purchase","Op_Leasing","Fin_Leasing","Subscription","Used_Car"]

    dict_content={
        "Private_User Purchase":User_Process,
        "Strategy Purchase":strategy_purchase,
        "Department Purchase":department_purchase
    }

    for p in perspective:
        for o in ownership:
            try:
                function=dict_content[f"{p} {o}"]
                template = f"<div class='{p} {o} content' style='display:none;'>{function()}</div>"
            except:
                template = f"<div class='{p} {o} content' style='display:none;'>{p} {o}</div>"

            print(template)





def merge():
    """
    merges 3 picture into 1 post together based on order in the folder


    1. loop over the folder to get 3 picture pair
    2. output
    :return:
    """

    counter=0
    prev=[]
    # copy image to other folder
    for (root, dirs, files) in os.walk("merge", topdown=True):
        for counter, file in enumerate(files):
            prev += [file]
            if len(prev)==3:
                #merge
                image1=Image.open(f'merge/{prev[0]}')
                image2=Image.open(f'merge/{prev[1]}')
                image3=Image.open(f'merge/{prev[2]}')
                new_image = Image.new('RGB', (3 * image1.size[0], image1.size[1]), (250, 250, 250))
                new_image.paste(image1, (image1.size[0]*0, 0))
                new_image.paste(image2, (image1.size[0]*1, 0))
                new_image.paste(image3, (image1.size[0]*2, 0))
                new_image.save(f"output/{counter}.jpg", "JPEG")
                prev = []
                print(f"created image {counter}")




if __name__ == '__main__':
    #by_Strategy()
    #by_Department()
    content_generator()



